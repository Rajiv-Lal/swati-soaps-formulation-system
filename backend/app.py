"""
SWATI SOAPS FORMULATION MANAGEMENT SYSTEM - BACKEND API
Flask REST API with SQLite
Version: 2.3 (Version Control Enhancements - Diff Generator & Change Reasons)
"""

from flask import Flask, request, jsonify
from ingredients_api import ingredients_bp
from sandbox_api import sandbox_bp
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from datetime import datetime, timedelta
import sqlite3
import json
import os
from functools import wraps
from dotenv import load_dotenv
import anthropic

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', 'jwt-secret-key-change-in-production')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)
app.config['DATABASE'] = os.environ.get('DATABASE_PATH', 'swati_soaps.db')

# Initialize extensions
app.register_blueprint(ingredients_bp)
app.register_blueprint(sandbox_bp)
CORS(app, origins=["http://localhost:3000", "http://165.22.222.87:3000", "http://165.22.222.87"], supports_credentials=True)
jwt = JWTManager(app)


def get_db():
    """Get database connection with production-safe settings"""
    conn = sqlite3.connect(app.config['DATABASE'], timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA busy_timeout = 30000')
    return conn


def dict_from_row(row):
    """Convert sqlite3.Row to dictionary"""
    return dict(zip(row.keys(), row)) if row else None


# ============================================================================
# VERSION CONTROL HELPERS
# ============================================================================

def generate_ingredient_diff(old_ingredients, new_ingredients):
    """
    Generate human-readable diff between old and new ingredient lists.
    Returns list of change descriptions like:
    - "Coconut Oil: 50% → 60%"
    - "Olive Oil removed"
    - "Shea Butter added at 5%"
    """
    changes = []

    # Create lookup dictionaries by ingredient_id
    old_map = {ing.get('ingredient_id'): ing for ing in old_ingredients}
    new_map = {ing.get('ingredient_id'): ing for ing in new_ingredients}

    # Check for modifications and removals
    for ing_id, old_ing in old_map.items():
        name = old_ing.get('ingredient_name', f'Ingredient #{ing_id}')
        old_pct = float(old_ing.get('percentage', 0))

        if ing_id in new_map:
            new_pct = float(new_map[ing_id].get('percentage', 0))
            if abs(old_pct - new_pct) > 0.01:
                changes.append(f"{name}: {old_pct:.1f}% → {new_pct:.1f}%")
        else:
            changes.append(f"{name} removed")

    # Check for additions
    for ing_id, new_ing in new_map.items():
        if ing_id not in old_map:
            name = new_ing.get('ingredient_name', f'Ingredient #{ing_id}')
            new_pct = float(new_ing.get('percentage', 0))
            changes.append(f"{name} added at {new_pct:.1f}%")

    return changes


def build_change_notes(user_notes, change_reasons, auto_diff):
    """
    Build comprehensive change notes combining:
    - User-provided notes
    - Selected change reasons (Price, Hardness, Perfume, Colour, Lather, Other)
    - Auto-generated ingredient diff
    """
    parts = []

    # Add change reasons if provided
    if change_reasons:
        reasons_text = "Reasons: " + ", ".join(change_reasons)
        parts.append(reasons_text)

    # Add user notes if provided
    if user_notes and user_notes.strip():
        parts.append(user_notes.strip())

    # Add auto-generated diff
    if auto_diff:
        diff_text = "Changes: " + "; ".join(auto_diff)
        parts.append(diff_text)

    return " | ".join(parts) if parts else "Updated formulation"


# Error handlers
@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Not found'}), 404


@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Internal server error'}), 500


# ============================================================================
# AUTHENTICATION ENDPOINTS
# ============================================================================

@app.route('/api/auth/login', methods=['POST'])
def login():
    """User login"""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'error': 'Email and password required'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM users WHERE email = ? AND is_active = 1', (email,))
        user = cursor.fetchone()
        
        if not user:
            conn.close()
            return jsonify({'error': 'Invalid credentials'}), 401
        
        cursor.execute('UPDATE users SET last_login = ? WHERE id = ?',
                      (datetime.now().isoformat(), user['id']))
        conn.commit()
        conn.close()
        
        access_token = create_access_token(identity=str(user['id']))
        
        return jsonify({
            'token': access_token,
            'user': {
                'id': user['id'],
                'email': user['email'],
                'name': user['full_name'],
                'role': user['role']
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/auth/me', methods=['GET'])
@jwt_required()
def get_current_user():
    """Get current user info"""
    try:
        user_id = get_jwt_identity()
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT id, email, full_name as name, role FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        conn.close()
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        return jsonify({'user': dict_from_row(user)}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# CATEGORY ENDPOINTS
# ============================================================================

@app.route('/api/categories', methods=['GET'])
@jwt_required()
def get_categories():
    """Get all categories with subcategories"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM categories ORDER BY display_order, name')
        categories = [dict_from_row(row) for row in cursor.fetchall()]
        
        for category in categories:
            cursor.execute(
                'SELECT * FROM subcategories WHERE category_id = ? ORDER BY display_order, name',
                (category['id'],)
            )
            category['subcategories'] = [dict_from_row(row) for row in cursor.fetchall()]
        
        conn.close()
        return jsonify({'categories': categories}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/subcategories/<int:category_id>', methods=['GET'])
@jwt_required()
def get_subcategories(category_id):
    """Get subcategories for a category"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute(
            'SELECT * FROM subcategories WHERE category_id = ? ORDER BY display_order, name',
            (category_id,)
        )
        subcategories = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'subcategories': subcategories}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# FORMULATION ENDPOINTS
# ============================================================================

@app.route('/api/formulations', methods=['GET'])
@jwt_required()
def get_formulations():
    """Get all formulations with filtering"""
    try:
        status = request.args.get('status')
        search = request.args.get('search')
        
        conn = get_db()
        cursor = conn.cursor()
        
        query = '''
            SELECT f.*, pt.name as product_type_name,
                   u.full_name as created_by_name,
                   COUNT(DISTINCT fi.ingredient_id) as ingredient_count
            FROM formulations f
            LEFT JOIN product_types pt ON f.product_type_id = pt.id
            LEFT JOIN users u ON f.created_by = u.id
            LEFT JOIN formulation_ingredients fi ON f.id = fi.formulation_id
            WHERE 1=1
        '''
        params = []
        
        if status:
            query += ' AND f.status = ?'
            params.append(status)
        
        if search:
            query += ' AND f.product_name LIKE ?'
            params.append(f'%{search}%')
        
        query += ' GROUP BY f.id ORDER BY f.created_at DESC'
        
        cursor.execute(query, params)
        formulations = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'formulations': formulations}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>', methods=['GET'])
@jwt_required()
def get_formulation(id):
    """Get single formulation with full details. Optionally load specific version."""
    try:
        version_id = request.args.get('version_id', type=int)

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT f.*, pt.name as product_type_name, u.full_name as created_by_name
            FROM formulations f
            LEFT JOIN product_types pt ON f.product_type_id = pt.id
            LEFT JOIN users u ON f.created_by = u.id
            WHERE f.id = ?
        ''', (id,))

        formulation = cursor.fetchone()

        if not formulation:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404

        result = dict_from_row(formulation)

        # If version_id provided, load from version snapshot
        if version_id:
            cursor.execute('''
                SELECT fv.*, u.full_name as created_by_name
                FROM formulation_versions fv
                LEFT JOIN users u ON fv.created_by = u.id
                WHERE fv.id = ? AND fv.formulation_id = ?
            ''', (version_id, id))
            version = cursor.fetchone()

            if version:
                result['loaded_version_id'] = version_id
                result['loaded_version_number'] = version['version_number']

                # Parse ingredients from snapshot
                snapshot = json.loads(version['ingredients_snapshot']) if version['ingredients_snapshot'] else {}
                snapshot_ingredients = snapshot.get('ingredients', [])

                # Get grammage from snapshot if available
                if 'grammage' in snapshot:
                    result['grammage'] = snapshot['grammage']
                if 'pack_count' in snapshot:
                    result['pack_count'] = snapshot['pack_count']

                # Enrich snapshot ingredients with current ingredient details
                enriched_ingredients = []
                for snap_ing in snapshot_ingredients:
                    cursor.execute('''
                        SELECT i.name as ingredient_name, i.inci_name, i.cas_number, i.landed_cost_net_gst,
                               c.name as category_name
                        FROM ingredients i
                        LEFT JOIN categories c ON i.category_id = c.id
                        WHERE i.id = ?
                    ''', (snap_ing['ingredient_id'],))
                    ing_details = cursor.fetchone()

                    enriched_ing = {
                        'ingredient_id': snap_ing['ingredient_id'],
                        'percentage': snap_ing['percentage'],
                        'quantity_grams': snap_ing.get('quantity_grams'),
                        'cost_per_piece': snap_ing.get('cost_per_piece'),
                        'ingredient_name': ing_details['ingredient_name'] if ing_details else f"Ingredient #{snap_ing['ingredient_id']}",
                        'landed_cost_net_gst': ing_details['landed_cost_net_gst'] if ing_details else 0,
                        'category_name': ing_details['category_name'] if ing_details else None,
                        'inci_name': ing_details['inci_name'] if ing_details else None,
                        'cas_number': ing_details['cas_number'] if ing_details else None
                    }
                    enriched_ingredients.append(enriched_ing)

                result['ingredients'] = enriched_ingredients
            else:
                # Version not found, fall back to current
                cursor.execute('''
                    SELECT fi.*, i.name as ingredient_name, i.inci_name, i.cas_number, i.landed_cost_net_gst,
                           c.name as category_name
                    FROM formulation_ingredients fi
                    JOIN ingredients i ON fi.ingredient_id = i.id
                    LEFT JOIN categories c ON i.category_id = c.id
                    WHERE fi.formulation_id = ?
                    ORDER BY i.name
                ''', (id,))
                result['ingredients'] = [dict_from_row(row) for row in cursor.fetchall()]
        else:
            # Get ingredients for current version
            cursor.execute('''
                SELECT fi.*, i.name as ingredient_name, i.inci_name, i.cas_number, i.landed_cost_net_gst,
                       c.name as category_name
                FROM formulation_ingredients fi
                JOIN ingredients i ON fi.ingredient_id = i.id
                LEFT JOIN categories c ON i.category_id = c.id
                WHERE fi.formulation_id = ?
                ORDER BY i.name
            ''', (id,))
            result['ingredients'] = [dict_from_row(row) for row in cursor.fetchall()]

        # Get benefits
        cursor.execute('''
            SELECT bc.id, bc.name
            FROM formulation_benefits fb
            JOIN benefit_categories bc ON fb.benefit_id = bc.id
            WHERE fb.formulation_id = ?
        ''', (id,))
        
        result['benefits'] = [dict_from_row(row) for row in cursor.fetchall()]
        
        # Get tags
        cursor.execute('''
            SELECT t.id, t.name, t.color
            FROM formulation_tags ft
            JOIN tags t ON ft.tag_id = t.id
            WHERE ft.formulation_id = ?
        ''', (id,))
        
        result['tags'] = [dict_from_row(row) for row in cursor.fetchall()]
        
        conn.close()
        return jsonify({'formulation': result}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations', methods=['POST'])
@jwt_required()
def create_formulation():
    """Create new formulation"""
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        # Validation
        if not data.get('product_name'):
            return jsonify({'error': 'Product name is required'}), 400
        
        if not data.get('grammage'):
            return jsonify({'error': 'Grammage is required'}), 400
        
        ingredients = data.get('ingredients', [])
        if not ingredients:
            return jsonify({'error': 'At least one ingredient is required'}), 400
        
        # Validate percentages sum to 100
        total_percentage = sum(float(ing.get('percentage', 0)) for ing in ingredients)
        if abs(total_percentage - 100.0) > 0.01:
            return jsonify({'error': f'Percentages must sum to 100% (currently {total_percentage:.2f}%)'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check for duplicate name
        cursor.execute('SELECT id FROM formulations WHERE product_name = ?', (data['product_name'],))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Product name already exists'}), 400
        
        # Calculate total cost and validate ingredients
        total_cost = 0
        grammage = float(data['grammage'])
        pack_count = int(data.get('pack_count', 1))
        regulatory_warnings = []  # Collect EU/US approval warnings
        
        for ing in ingredients:
            cursor.execute('''
                SELECT i.name, i.landed_cost_net_gst, i.usage_rate_min, i.usage_rate_max,
                       r.eu_approved, r.us_approved
                FROM ingredients i
                LEFT JOIN ingredient_regulatory r ON i.id = r.ingredient_id
                WHERE i.id = ?
            ''', (ing['ingredient_id'],))
            ingredient = cursor.fetchone()
            
            if not ingredient:
                conn.close()
                return jsonify({'error': f'Ingredient ID {ing["ingredient_id"]} not found'}), 400
            
            percentage = float(ing['percentage'])
            
            # Validate usage rate limits (WARNING with override option)
            override_limits = data.get('override_usage_limits', False)
            
            if ingredient['usage_rate_max'] is not None and percentage > float(ingredient['usage_rate_max']):
                if not override_limits:
                    regulatory_warnings.append({
                        'type': 'usage_rate_max',
                        'ingredient': ingredient['name'],
                        'limit': float(ingredient['usage_rate_max']),
                        'actual': percentage,
                        'message': f'"{ingredient["name"]}" exceeds maximum usage rate ({ingredient["usage_rate_max"]}%). You specified {percentage}%'
                    })
            
            if ingredient['usage_rate_min'] is not None and percentage < float(ingredient['usage_rate_min']):
                if not override_limits:
                    regulatory_warnings.append({
                        'type': 'usage_rate_min',
                        'ingredient': ingredient['name'],
                        'limit': float(ingredient['usage_rate_min']),
                        'actual': percentage,
                        'message': f'"{ingredient["name"]}" is below minimum usage rate ({ingredient["usage_rate_min"]}%). You specified {percentage}%'
                    })
            
            # Check regulatory approval (ADVISORY WARNING - not a hard stop)
            eu_approved = ingredient['eu_approved'] if ingredient['eu_approved'] is not None else 1
            us_approved = ingredient['us_approved'] if ingredient['us_approved'] is not None else 1
            
            if not eu_approved:
                regulatory_warnings.append(f'"{ingredient["name"]}" is not EU approved')
            
            if not us_approved:
                regulatory_warnings.append(f'"{ingredient["name"]}" is not US approved')
            
            # Calculate costs
            cost_per_kg = float(ingredient['landed_cost_net_gst'] or 0)
            quantity_grams = (grammage * percentage) / 100
            ing['quantity_grams'] = quantity_grams
            ing['cost_per_piece'] = (quantity_grams / 1000) * cost_per_kg
            total_cost += ing['cost_per_piece']
        
        # Insert formulation
        cursor.execute('''
            INSERT INTO formulations (
                product_name, product_type_id, current_version, grammage,
                pack_count, status, total_cost_per_piece,
                notes, created_at, updated_at, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['product_name'],
            data.get('product_type_id'),
            'v1.0',
            grammage,
            pack_count,
            data.get('status', 'draft'),
            round(total_cost, 4),
            data.get('notes'),
            datetime.now().isoformat(),
            datetime.now().isoformat(),
            user_id
        ))
        
        formulation_id = cursor.lastrowid
        
        # Create initial version
        snapshot = {
            'grammage': grammage,
            'pack_count': pack_count,
            'ingredients': ingredients
        }
        
        cursor.execute('''
            INSERT INTO formulation_versions (
                formulation_id, version_number, created_at, created_by,
                change_notes, ingredients_snapshot, cost_snapshot
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            formulation_id,
            'v1.0',
            datetime.now().isoformat(),
            user_id,
            'Initial version',
            json.dumps(snapshot),
            round(total_cost, 4)
        ))
        
        version_id = cursor.lastrowid
        
        # Insert ingredients with version_id
        for ing in ingredients:
            cursor.execute('''
                INSERT INTO formulation_ingredients (
                    formulation_id, version_id, ingredient_id, percentage,
                    quantity_grams, cost_per_piece
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                formulation_id,
                version_id,
                ing['ingredient_id'],
                ing['percentage'],
                ing['quantity_grams'],
                ing['cost_per_piece']
            ))
        
        # Add benefits
        for benefit_id in data.get('benefits', []):
            cursor.execute(
                'INSERT INTO formulation_benefits (formulation_id, benefit_id) VALUES (?, ?)',
                (formulation_id, benefit_id)
            )
        
        # Add tags
        for tag_id in data.get('tags', []):
            cursor.execute(
                'INSERT INTO formulation_tags (formulation_id, tag_id, tagged_by) VALUES (?, ?, ?)',
                (formulation_id, tag_id, user_id)
            )
        
        conn.commit()
        conn.close()
        
        response = {
            'message': 'Formulation created successfully',
            'formulation_id': formulation_id,
            'total_cost': round(total_cost, 4)
        }
        
        # Include regulatory warnings if any
        if regulatory_warnings:
            response['warnings'] = regulatory_warnings
        
        return jsonify(response), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>', methods=['PUT'])
@jwt_required()
def update_formulation(id):
    """Update formulation - creates new version if ingredients changed"""
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if exists
        cursor.execute('SELECT * FROM formulations WHERE id = ?', (id,))
        existing = cursor.fetchone()
        
        if not existing:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404
        
        ingredients = data.get('ingredients', [])
        
        # Validate percentages if ingredients provided
        if ingredients:
            total_percentage = sum(float(ing.get('percentage', 0)) for ing in ingredients)
            if abs(total_percentage - 100.0) > 0.01:
                conn.close()
                return jsonify({'error': f'Percentages must sum to 100% (currently {total_percentage:.2f}%)'}), 400
        
        # Only create new version if user explicitly requests it
        create_new_version = data.get('create_new_version', False)

        # Calculate costs and validate ingredients
        total_cost = 0
        grammage = float(data.get('grammage', existing['grammage']))
        pack_count = int(data.get('pack_count', existing['pack_count']))
        regulatory_warnings = []  # Collect EU/US approval warnings
        
        if ingredients:
            for ing in ingredients:
                cursor.execute('''
                    SELECT i.name, i.landed_cost_net_gst, i.usage_rate_min, i.usage_rate_max,
                           r.eu_approved, r.us_approved
                    FROM ingredients i
                    LEFT JOIN ingredient_regulatory r ON i.id = r.ingredient_id
                    WHERE i.id = ?
                ''', (ing['ingredient_id'],))
                ingredient = cursor.fetchone()
                
                if not ingredient:
                    conn.close()
                    return jsonify({'error': f'Ingredient ID {ing["ingredient_id"]} not found'}), 400
                
                percentage = float(ing['percentage'])
                
                # Validate usage rate limits (WARNING with override option)
                override_limits = data.get('override_usage_limits', False)
                
                if ingredient['usage_rate_max'] is not None and percentage > float(ingredient['usage_rate_max']):
                    if not override_limits:
                        regulatory_warnings.append({
                            'type': 'usage_rate_max',
                            'ingredient': ingredient['name'],
                            'limit': float(ingredient['usage_rate_max']),
                            'actual': percentage,
                            'message': f'"{ingredient["name"]}" exceeds maximum usage rate ({ingredient["usage_rate_max"]}%). You specified {percentage}%'
                        })
                
                if ingredient['usage_rate_min'] is not None and percentage < float(ingredient['usage_rate_min']):
                    if not override_limits:
                        regulatory_warnings.append({
                            'type': 'usage_rate_min',
                            'ingredient': ingredient['name'],
                            'limit': float(ingredient['usage_rate_min']),
                            'actual': percentage,
                            'message': f'"{ingredient["name"]}" is below minimum usage rate ({ingredient["usage_rate_min"]}%). You specified {percentage}%'
                        })
                
                # Check regulatory approval (ADVISORY WARNING - not a hard stop)
                eu_approved = ingredient['eu_approved'] if ingredient['eu_approved'] is not None else 1
                us_approved = ingredient['us_approved'] if ingredient['us_approved'] is not None else 1
                
                if not eu_approved:
                    regulatory_warnings.append(f'"{ingredient["name"]}" is not EU approved')
                
                if not us_approved:
                    regulatory_warnings.append(f'"{ingredient["name"]}" is not US approved')
                
                # Calculate costs
                cost_per_kg = float(ingredient['landed_cost_net_gst'] or 0)
                quantity_grams = (grammage * percentage) / 100
                ing['quantity_grams'] = quantity_grams
                ing['cost_per_piece'] = (quantity_grams / 1000) * cost_per_kg
                total_cost += ing['cost_per_piece']
        else:
            total_cost = existing['total_cost_per_piece'] or 0
        
        # Determine new version number
        new_version = existing['current_version']
        version_id = None
        
        if create_new_version and ingredients:
            # Increment version
            version_parts = existing['current_version'].replace('v', '').split('.')
            major = int(version_parts[0])
            minor = int(version_parts[1])

            minor += 1
            if minor >= 10:
                major += 1
                minor = 0

            new_version = f'v{major}.{minor}'

            # Get old ingredients for diff generation
            cursor.execute('''
                SELECT fi.ingredient_id, fi.percentage, i.name as ingredient_name
                FROM formulation_ingredients fi
                JOIN ingredients i ON fi.ingredient_id = i.id
                WHERE fi.formulation_id = ?
            ''', (id,))
            old_ingredients = [dict(row) for row in cursor.fetchall()]

            # Enrich new ingredients with names for diff
            new_ingredients_with_names = []
            for ing in ingredients:
                cursor.execute('SELECT name FROM ingredients WHERE id = ?', (ing['ingredient_id'],))
                ing_row = cursor.fetchone()
                new_ingredients_with_names.append({
                    'ingredient_id': ing['ingredient_id'],
                    'percentage': ing['percentage'],
                    'ingredient_name': ing_row['name'] if ing_row else f"Ingredient #{ing['ingredient_id']}"
                })

            # Generate auto-diff
            auto_diff = generate_ingredient_diff(old_ingredients, new_ingredients_with_names)

            # Get change reasons from request
            change_reasons = data.get('change_reasons', [])
            user_notes = data.get('version_notes', '')

            # Build comprehensive change notes
            final_notes = build_change_notes(user_notes, change_reasons, auto_diff)

            # Create version snapshot
            snapshot = {
                'grammage': grammage,
                'pack_count': pack_count,
                'ingredients': ingredients
            }

            cursor.execute('''
                INSERT INTO formulation_versions (
                    formulation_id, version_number, created_at, created_by,
                    change_notes, ingredients_snapshot, cost_snapshot
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                id,
                new_version,
                datetime.now().isoformat(),
                user_id,
                final_notes,
                json.dumps(snapshot),
                round(total_cost, 4)
            ))

            version_id = cursor.lastrowid
        
        # Update formulation record
        cursor.execute('''
            UPDATE formulations SET
                product_name = ?, product_type_id = ?, current_version = ?,
                grammage = ?, pack_count = ?, status = ?,
                total_cost_per_piece = ?,
                notes = ?, updated_at = ?
            WHERE id = ?
        ''', (
            data.get('product_name', existing['product_name']),
            data.get('product_type_id', existing['product_type_id']),
            new_version,
            grammage,
            pack_count,
            data.get('status', existing['status']),
            round(total_cost, 4),
            data.get('notes', existing['notes']),
            datetime.now().isoformat(),
            id
        ))
        
        # Update ingredients if provided
        if ingredients:
            # Delete old ingredients
            cursor.execute('DELETE FROM formulation_ingredients WHERE formulation_id = ?', (id,))
            
            # Get version_id if not creating new version
            if not version_id:
                cursor.execute('''
                    SELECT id FROM formulation_versions
                    WHERE formulation_id = ?
                    ORDER BY created_at DESC LIMIT 1
                ''', (id,))
                version_row = cursor.fetchone()
                version_id = version_row['id'] if version_row else None
            
            # Insert new ingredients
            for ing in ingredients:
                cursor.execute('''
                    INSERT INTO formulation_ingredients (
                        formulation_id, version_id, ingredient_id, percentage,
                        quantity_grams, cost_per_piece
                    ) VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    id,
                    version_id,
                    ing['ingredient_id'],
                    ing['percentage'],
                    ing['quantity_grams'],
                    ing['cost_per_piece']
                ))
        
        # Update benefits if provided
        if 'benefits' in data:
            cursor.execute('DELETE FROM formulation_benefits WHERE formulation_id = ?', (id,))
            for benefit_id in data['benefits']:
                cursor.execute(
                    'INSERT INTO formulation_benefits (formulation_id, benefit_id) VALUES (?, ?)',
                    (id, benefit_id)
                )
        
        # Update tags if provided
        if 'tags' in data:
            cursor.execute('DELETE FROM formulation_tags WHERE formulation_id = ?', (id,))
            for tag_id in data['tags']:
                cursor.execute(
                    'INSERT INTO formulation_tags (formulation_id, tag_id, tagged_by) VALUES (?, ?, ?)',
                    (id, tag_id, user_id)
                )
        
        conn.commit()
        conn.close()
        
        response = {
            'message': 'Formulation updated successfully',
            'new_version': new_version if create_new_version else None,
            'total_cost': round(total_cost, 4)
        }
        
        # Include regulatory warnings if any
        if regulatory_warnings:
            response['warnings'] = regulatory_warnings
        
        return jsonify(response), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>/duplicate', methods=['POST'])
@jwt_required()
def duplicate_formulation(id):
    """Duplicate an existing formulation"""
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Get original formulation
        cursor.execute('SELECT * FROM formulations WHERE id = ?', (id,))
        original = cursor.fetchone()
        
        if not original:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404
        
        # Create new name
        new_name = data.get('new_name', f"{original['product_name']} (Copy)")
        
        # Check for duplicate name
        cursor.execute('SELECT id FROM formulations WHERE product_name = ?', (new_name,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Product name already exists'}), 400
        
        # Insert new formulation
        cursor.execute('''
            INSERT INTO formulations (
                product_name, product_type_id, current_version, grammage,
                pack_count, status, total_cost_per_piece,
                notes, created_at, updated_at, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            new_name,
            original['product_type_id'],
            'v1.0',
            original['grammage'],
            original['pack_count'],
            'draft',
            original['total_cost_per_piece'],
            f"Duplicated from {original['product_name']}",
            datetime.now().isoformat(),
            datetime.now().isoformat(),
            user_id
        ))
        
        new_formulation_id = cursor.lastrowid
        
        # Get original ingredients
        cursor.execute('''
            SELECT * FROM formulation_ingredients
            WHERE formulation_id = ?
        ''', (id,))
        
        original_ingredients = [dict_from_row(row) for row in cursor.fetchall()]
        
        # Create snapshot for version
        snapshot = {
            'grammage': original['grammage'],
            'pack_count': original['pack_count'],
            'ingredients': original_ingredients
        }
        
        # Create initial version for duplicate
        cursor.execute('''
            INSERT INTO formulation_versions (
                formulation_id, version_number, created_at, created_by,
                change_notes, ingredients_snapshot, cost_snapshot
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            new_formulation_id,
            'v1.0',
            datetime.now().isoformat(),
            user_id,
            f'Duplicated from {original["product_name"]}',
            json.dumps(snapshot),
            original['total_cost_per_piece']
        ))
        
        new_version_id = cursor.lastrowid
        
        # Copy ingredients
        for ing in original_ingredients:
            cursor.execute('''
                INSERT INTO formulation_ingredients (
                    formulation_id, version_id, ingredient_id, percentage,
                    quantity_grams, cost_per_piece
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                new_formulation_id,
                new_version_id,
                ing['ingredient_id'],
                ing['percentage'],
                ing['quantity_grams'],
                ing['cost_per_piece']
            ))
        
        # Copy benefits
        cursor.execute('''
            INSERT INTO formulation_benefits (formulation_id, benefit_id)
            SELECT ?, benefit_id FROM formulation_benefits WHERE formulation_id = ?
        ''', (new_formulation_id, id))
        
        # Copy tags
        cursor.execute('''
            INSERT INTO formulation_tags (formulation_id, tag_id, tagged_by)
            SELECT ?, tag_id, ? FROM formulation_tags WHERE formulation_id = ?
        ''', (new_formulation_id, user_id, id))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': 'Formulation duplicated successfully',
            'formulation_id': new_formulation_id
        }), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_formulation(id):
    """Delete formulation"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT product_name FROM formulations WHERE id = ?', (id,))
        formulation = cursor.fetchone()
        
        if not formulation:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404
        
        cursor.execute('DELETE FROM formulations WHERE id = ?', (id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Formulation deleted successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>/versions/<int:version_id>', methods=['DELETE'])
@jwt_required()
def delete_formulation_version(id, version_id):
    """Delete a specific version of a formulation.
    - Cannot delete if it's the only version
    - Updates current_version if the deleted version was current
    """
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Check formulation exists
        cursor.execute('SELECT id, current_version FROM formulations WHERE id = ?', (id,))
        formulation = cursor.fetchone()
        if not formulation:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404

        # Check version exists
        cursor.execute('SELECT id, version_number FROM formulation_versions WHERE id = ? AND formulation_id = ?',
                      (version_id, id))
        version = cursor.fetchone()
        if not version:
            conn.close()
            return jsonify({'error': 'Version not found'}), 404

        # Count total versions
        cursor.execute('SELECT COUNT(*) as count FROM formulation_versions WHERE formulation_id = ?', (id,))
        version_count = cursor.fetchone()['count']

        # Cannot delete if only one version
        if version_count <= 1:
            conn.close()
            return jsonify({'error': 'Cannot delete the only version. Delete the entire formulation instead.'}), 400

        deleted_version_number = version['version_number']
        current_version = formulation['current_version']

        # Delete the version
        cursor.execute('DELETE FROM formulation_versions WHERE id = ?', (version_id,))

        # If deleted version was the current one, update to the latest remaining version
        if deleted_version_number == current_version:
            cursor.execute('''
                SELECT version_number FROM formulation_versions
                WHERE formulation_id = ?
                ORDER BY created_at DESC
                LIMIT 1
            ''', (id,))
            latest = cursor.fetchone()
            if latest:
                cursor.execute('UPDATE formulations SET current_version = ? WHERE id = ?',
                              (latest['version_number'], id))

        conn.commit()
        conn.close()

        return jsonify({
            'message': f'Version {deleted_version_number} deleted successfully',
            'deleted_version': deleted_version_number
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# VERSION CONTROL ENDPOINTS
# ============================================================================

@app.route('/api/formulations/<int:id>/versions', methods=['GET'])
@jwt_required()
def get_formulation_versions(id):
    """Get all versions of a formulation with ingredient names"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT fv.*, u.full_name as created_by_name
            FROM formulation_versions fv
            LEFT JOIN users u ON fv.created_by = u.id
            WHERE fv.formulation_id = ?
            ORDER BY fv.created_at DESC
        ''', (id,))
        
        versions = [dict_from_row(row) for row in cursor.fetchall()]
        
        # Collect all ingredient IDs across all versions
        all_ingredient_ids = set()
        for version in versions:
            if version.get('ingredients_snapshot'):
                snapshot = json.loads(version['ingredients_snapshot'])
                for ing in snapshot.get('ingredients', []):
                    all_ingredient_ids.add(ing['ingredient_id'])
        
        # Fetch all ingredient names in one query
        ingredient_names = {}
        if all_ingredient_ids:
            placeholders = ','.join('?' * len(all_ingredient_ids))
            cursor.execute(f'SELECT id, name FROM ingredients WHERE id IN ({placeholders})', 
                          list(all_ingredient_ids))
            ingredient_names = {row['id']: row['name'] for row in cursor.fetchall()}
        
        # Enrich each version with ingredient names
        for version in versions:
            if version.get('ingredients_snapshot'):
                snapshot = json.loads(version['ingredients_snapshot'])
                for ing in snapshot.get('ingredients', []):
                    ing['name'] = ingredient_names.get(ing['ingredient_id'], f"Ingredient #{ing['ingredient_id']}")
                version['ingredients_snapshot'] = snapshot
        
        conn.close()
        return jsonify({'versions': versions}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>/versions/compare', methods=['GET'])
@jwt_required()
def compare_versions(id):
    """Compare two versions of a formulation"""
    try:
        v1 = request.args.get('v1')
        v2 = request.args.get('v2')
        
        if not v1 or not v2:
            return jsonify({'error': 'Both v1 and v2 parameters required'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Get version 1
        cursor.execute('''
            SELECT * FROM formulation_versions
            WHERE formulation_id = ? AND version_number = ?
        ''', (id, v1))
        version1 = cursor.fetchone()
        
        # Get version 2
        cursor.execute('''
            SELECT * FROM formulation_versions
            WHERE formulation_id = ? AND version_number = ?
        ''', (id, v2))
        version2 = cursor.fetchone()
        
        if not version1 or not version2:
            conn.close()
            return jsonify({'error': 'Version not found'}), 404
        
        v1_dict = dict_from_row(version1)
        v2_dict = dict_from_row(version2)
        
        # Parse snapshots
        v1_snapshot = json.loads(v1_dict['ingredients_snapshot']) if v1_dict.get('ingredients_snapshot') else {}
        v2_snapshot = json.loads(v2_dict['ingredients_snapshot']) if v2_dict.get('ingredients_snapshot') else {}
        
        v1_ingredients = v1_snapshot.get('ingredients', [])
        v2_ingredients = v2_snapshot.get('ingredients', [])
        
        v1_ing_map = {ing['ingredient_id']: ing for ing in v1_ingredients}
        v2_ing_map = {ing['ingredient_id']: ing for ing in v2_ingredients}
        
        added = [ing_id for ing_id in v2_ing_map if ing_id not in v1_ing_map]
        removed = [ing_id for ing_id in v1_ing_map if ing_id not in v2_ing_map]
        modified = []
        
        for ing_id in v1_ing_map:
            if ing_id in v2_ing_map:
                if v1_ing_map[ing_id]['percentage'] != v2_ing_map[ing_id]['percentage']:
                    modified.append({
                        'ingredient_id': ing_id,
                        'old_percentage': v1_ing_map[ing_id]['percentage'],
                        'new_percentage': v2_ing_map[ing_id]['percentage']
                    })
        
        # Get ingredient names
        ing_ids = list(set(added + removed + [m['ingredient_id'] for m in modified]))
        ing_names = {}
        if ing_ids:
            placeholders = ','.join('?' * len(ing_ids))
            cursor.execute(f'SELECT id, name FROM ingredients WHERE id IN ({placeholders})', ing_ids)
            ing_names = {row['id']: row['name'] for row in cursor.fetchall()}
        
        conn.close()
        
        v1_dict['ingredients_snapshot'] = v1_snapshot
        v2_dict['ingredients_snapshot'] = v2_snapshot
        
        return jsonify({
            'version1': v1_dict,
            'version2': v2_dict,
            'differences': {
                'added': [{'ingredient_id': iid, 'name': ing_names.get(iid)} for iid in added],
                'removed': [{'ingredient_id': iid, 'name': ing_names.get(iid)} for iid in removed],
                'modified': [{**m, 'name': ing_names.get(m['ingredient_id'])} for m in modified],
                'cost_change': (v2_dict['cost_snapshot'] or 0) - (v1_dict['cost_snapshot'] or 0)
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>/benefits', methods=['GET'])
@jwt_required()
def get_formulation_benefits(id):
    """Get marketing benefits for all ingredients in a formulation"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                i.id as ingredient_id,
                i.name as ingredient_name,
                i.inci_name,
                fi.percentage,
                im.benefits,
                im.applications
            FROM formulation_ingredients fi
            JOIN ingredients i ON fi.ingredient_id = i.id
            LEFT JOIN ingredient_marketing im ON i.id = im.ingredient_id
            WHERE fi.formulation_id = ?
            ORDER BY fi.percentage DESC
        ''', (id,))
        
        ingredients_benefits = []
        benefit_counts = {}  # Track frequency of each benefit
        application_counts = {}  # Track frequency of each application

        for row in cursor.fetchall():
            row_dict = dict_from_row(row)
            ingredients_benefits.append(row_dict)

            if row_dict.get('benefits'):
                for benefit in row_dict['benefits'].split(','):
                    benefit = benefit.strip().lower()  # Normalize to lowercase
                    if benefit:
                        # Capitalize for display
                        display_benefit = benefit.capitalize()
                        benefit_counts[display_benefit] = benefit_counts.get(display_benefit, 0) + 1

            if row_dict.get('applications'):
                for app in row_dict['applications'].split(','):
                    app = app.strip().lower()
                    if app:
                        display_app = app.capitalize()
                        application_counts[display_app] = application_counts.get(display_app, 0) + 1

        conn.close()

        # Sort by frequency (highest first), then alphabetically
        sorted_benefits = sorted(
            benefit_counts.items(),
            key=lambda x: (-x[1], x[0])  # -count for descending, then name
        )
        sorted_applications = sorted(
            application_counts.items(),
            key=lambda x: (-x[1], x[0])
        )

        return jsonify({
            'ingredients_benefits': ingredients_benefits,
            'consolidated_benefits': [
                {'benefit': b, 'frequency': c} for b, c in sorted_benefits
            ],
            'consolidated_applications': [
                {'application': a, 'frequency': c} for a, c in sorted_applications
            ],
            'total_ingredients': len(ingredients_benefits)
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>/marketing-benefits', methods=['GET'])
@jwt_required()
def get_marketing_benefits(id):
    """Generate marketing benefit statements using Claude API"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Get formulation name
        cursor.execute('SELECT product_name FROM formulations WHERE id = ?', (id,))
        formulation = cursor.fetchone()
        if not formulation:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404

        product_name = formulation['product_name']

        # Get ingredients with their benefits
        cursor.execute('''
            SELECT
                i.name as ingredient_name,
                fi.percentage,
                im.benefits
            FROM formulation_ingredients fi
            JOIN ingredients i ON fi.ingredient_id = i.id
            LEFT JOIN ingredient_marketing im ON i.id = im.ingredient_id
            WHERE fi.formulation_id = ?
            ORDER BY fi.percentage DESC
        ''', (id,))

        # Collect benefits with frequency
        benefit_counts = {}
        ingredients_list = []

        for row in cursor.fetchall():
            row_dict = dict_from_row(row)
            ingredients_list.append(row_dict['ingredient_name'])

            if row_dict.get('benefits'):
                for benefit in row_dict['benefits'].split(','):
                    benefit = benefit.strip().lower()
                    if benefit:
                        display_benefit = benefit.capitalize()
                        benefit_counts[display_benefit] = benefit_counts.get(display_benefit, 0) + 1

        conn.close()

        if not benefit_counts:
            return jsonify({
                'marketing_statements': [],
                'message': 'No benefits data available for ingredients'
            }), 200

        # Sort benefits by frequency
        sorted_benefits = sorted(
            benefit_counts.items(),
            key=lambda x: (-x[1], x[0])
        )

        # Prepare benefits text for Claude
        benefits_text = "\n".join([
            f"- {benefit} (appears in {count} ingredient{'s' if count > 1 else ''})"
            for benefit, count in sorted_benefits
        ])

        # Check for API key
        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            # Return raw benefits if no API key
            return jsonify({
                'marketing_statements': [b for b, c in sorted_benefits[:10]],
                'raw_benefits': [{'benefit': b, 'frequency': c} for b, c in sorted_benefits],
                'message': 'ANTHROPIC_API_KEY not set - returning raw benefits'
            }), 200

        # Call Claude API
        client = anthropic.Anthropic(api_key=api_key)

        prompt = f"""Based on the following ingredient benefits for a soap product, generate 5-7 clear, factual marketing benefit statements.

Product: {product_name}
Key Ingredients: {', '.join(ingredients_list[:8])}

Ingredient Benefits (ranked by frequency - benefits appearing in more ingredients are more significant):
{benefits_text}

Guidelines:
- Write clear, factual statements without hyperbole
- Avoid words like "revolutionary", "miracle", "amazing", "best"
- Focus on what the product actually does for skin
- Prioritize benefits that appear in multiple ingredients
- Each statement should be 1 sentence, under 15 words
- Use simple language a consumer would understand

Return ONLY the benefit statements, one per line, no numbering or bullets."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=500,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )

        # Parse response
        response_text = message.content[0].text
        statements = [s.strip() for s in response_text.strip().split('\n') if s.strip()]

        return jsonify({
            'marketing_statements': statements,
            'raw_benefits': [{'benefit': b, 'frequency': c} for b, c in sorted_benefits],
            'total_ingredients': len(ingredients_list),
            'product_name': product_name
        }), 200

    except anthropic.APIError as e:
        return jsonify({
            'error': f'Claude API error: {str(e)}',
            'raw_benefits': [{'benefit': b, 'frequency': c} for b, c in sorted_benefits] if 'sorted_benefits' in locals() else []
        }), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>/versions/<int:version_id>/restore', methods=['POST'])
@jwt_required()
def restore_version(id, version_id):
    """Restore a previous version (creates new version)"""
    try:
        user_id = get_jwt_identity()
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Get the version to restore
        cursor.execute('''
            SELECT * FROM formulation_versions
            WHERE id = ? AND formulation_id = ?
        ''', (version_id, id))
        
        old_version = cursor.fetchone()
        
        if not old_version:
            conn.close()
            return jsonify({'error': 'Version not found'}), 404
        
        # Get current formulation
        cursor.execute('SELECT * FROM formulations WHERE id = ?', (id,))
        formulation = cursor.fetchone()
        
        # Increment version
        version_parts = formulation['current_version'].replace('v', '').split('.')
        major = int(version_parts[0])
        minor = int(version_parts[1]) + 1
        
        if minor >= 10:
            major += 1
            minor = 0
        
        new_version = f'v{major}.{minor}'
        
        # Parse old snapshot
        old_snapshot = json.loads(old_version['ingredients_snapshot']) if old_version['ingredients_snapshot'] else {}
        ingredients = old_snapshot.get('ingredients', [])
        grammage = old_snapshot.get('grammage', formulation['grammage'])
        pack_count = old_snapshot.get('pack_count', formulation['pack_count'])
        
        # Create new version from old snapshot
        cursor.execute('''
            INSERT INTO formulation_versions (
                formulation_id, version_number, created_at, created_by,
                change_notes, ingredients_snapshot, cost_snapshot
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            id,
            new_version,
            datetime.now().isoformat(),
            user_id,
            f"Restored from {old_version['version_number']}",
            old_version['ingredients_snapshot'],
            old_version['cost_snapshot']
        ))
        
        new_version_id = cursor.lastrowid
        
        # Delete current ingredients
        cursor.execute('DELETE FROM formulation_ingredients WHERE formulation_id = ?', (id,))
        
        # Insert restored ingredients
        total_cost = 0
        for ing in ingredients:
            # Recalculate cost with current prices
            cursor.execute(
                'SELECT landed_cost_net_gst FROM ingredients WHERE id = ?',
                (ing['ingredient_id'],)
            )
            ingredient = cursor.fetchone()
            
            if ingredient:
                cost_per_kg = float(ingredient['landed_cost_net_gst'] or 0)
                percentage = float(ing['percentage'])
                quantity_grams = (grammage * percentage) / 100
                cost_per_piece = (quantity_grams / 1000) * cost_per_kg
                total_cost += cost_per_piece
                
                cursor.execute('''
                    INSERT INTO formulation_ingredients (
                        formulation_id, version_id, ingredient_id, percentage,
                        quantity_grams, cost_per_piece
                    ) VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    id,
                    new_version_id,
                    ing['ingredient_id'],
                    percentage,
                    quantity_grams,
                    cost_per_piece
                ))
        
        # Update formulation
        cursor.execute('''
            UPDATE formulations SET
                current_version = ?,
                grammage = ?,
                pack_count = ?,
                total_cost_per_piece = ?,
                updated_at = ?
            WHERE id = ?
        ''', (
            new_version,
            grammage,
            pack_count,
            round(total_cost, 4),
            datetime.now().isoformat(),
            id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'Restored to version {old_version["version_number"]} as {new_version}',
            'new_version': new_version
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# TEST RESULTS ENDPOINTS
# ============================================================================

@app.route('/api/formulations/<int:id>/tests', methods=['GET'])
@jwt_required()
def get_test_results(id):
    """Get all test results for a formulation"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT tr.*, u.full_name as tested_by_name, fv.version_number
            FROM test_results tr
            LEFT JOIN users u ON tr.tested_by = u.id
            LEFT JOIN formulation_versions fv ON tr.version_id = fv.id
            WHERE tr.formulation_id = ?
            ORDER BY tr.test_date DESC
        ''', (id,))
        
        tests = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'tests': tests}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>/tests', methods=['POST'])
@jwt_required()
def create_test_result(id):
    """Add test result for a formulation"""
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT id FROM formulations WHERE id = ?', (id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404
        
        cursor.execute('''
            INSERT INTO test_results (
                formulation_id, version_id, test_date, tested_by,
                hardness_value, hardness_method,
                lather_quality, lather_quantity, lather_stability,
                notes, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            id,
            data.get('version_id'),
            data.get('test_date', datetime.now().date().isoformat()),
            user_id,
            data.get('hardness_value'),
            data.get('hardness_method'),
            data.get('lather_quality'),
            data.get('lather_quantity'),
            data.get('lather_stability'),
            data.get('notes'),
            datetime.now().isoformat()
        ))
        
        test_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': 'Test result added successfully',
            'test_id': test_id
        }), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tests/<int:test_id>', methods=['PUT'])
@jwt_required()
def update_test_result(test_id):
    """Update test result"""
    try:
        data = request.get_json()
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT id FROM test_results WHERE id = ?', (test_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Test result not found'}), 404
        
        cursor.execute('''
            UPDATE test_results SET
                test_date = ?,
                hardness_value = ?,
                hardness_method = ?,
                lather_quality = ?,
                lather_quantity = ?,
                lather_stability = ?,
                notes = ?
            WHERE id = ?
        ''', (
            data.get('test_date'),
            data.get('hardness_value'),
            data.get('hardness_method'),
            data.get('lather_quality'),
            data.get('lather_quantity'),
            data.get('lather_stability'),
            data.get('notes'),
            test_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Test result updated successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tests/<int:test_id>', methods=['DELETE'])
@jwt_required()
def delete_test_result(test_id):
    """Delete test result"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM test_results WHERE id = ?', (test_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'Test result not found'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Test result deleted successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# BOM GENERATION ENDPOINT
# ============================================================================

@app.route('/api/formulations/<int:id>/bom/generate', methods=['POST'])
@jwt_required()
def generate_bom(id):
    """Generate Bill of Materials with wastage calculation and version support"""
    try:
        data = request.get_json() or {}
        
        target_quantity = int(data.get('quantity', 1000))
        pack_count = int(data.get('pack_count', 1))
        wastage_percent = float(data.get('wastage_percent', 2.0))
        version_id = data.get('version_id')  # Optional: specific version
        batch_unit = data.get('batch_unit', 'kg')  # 'kg' or 'tonnes'
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Get formulation with optional specific version
        if version_id:
            cursor.execute('''
                SELECT f.*, fv.version_number, fv.id as version_id
                FROM formulations f
                JOIN formulation_versions fv ON f.id = fv.formulation_id
                WHERE f.id = ? AND fv.id = ?
            ''', (id, version_id))
        else:
            cursor.execute('''
                SELECT f.*, fv.version_number, fv.id as version_id
                FROM formulations f
                LEFT JOIN formulation_versions fv ON f.id = fv.formulation_id
                WHERE f.id = ?
                ORDER BY fv.created_at DESC
                LIMIT 1
            ''', (id,))
        
        formulation = cursor.fetchone()
        
        if not formulation:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404
        
        # Get ingredients for specific version if version_id provided
        if version_id and formulation['version_id']:
            cursor.execute('''
                SELECT fi.*, i.name as ingredient_name, i.inci_name, i.cas_number, i.landed_cost_net_gst,
                       c.name as category_name
                FROM formulation_ingredients fi
                JOIN ingredients i ON fi.ingredient_id = i.id
                LEFT JOIN categories c ON i.category_id = c.id
                WHERE fi.formulation_id = ? AND fi.version_id = ?
            ''', (id, version_id))
        else:
            cursor.execute('''
                SELECT fi.*, i.name as ingredient_name, i.inci_name, i.cas_number, i.landed_cost_net_gst,
                       c.name as category_name
                FROM formulation_ingredients fi
                JOIN ingredients i ON fi.ingredient_id = i.id
                LEFT JOIN categories c ON i.category_id = c.id
                WHERE fi.formulation_id = ?
            ''', (id,))
        
        ingredients = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        grammage = float(formulation['grammage'])
        total_pieces = target_quantity * pack_count
        
        bom_items = []
        total_cost = 0
        
        for ing in ingredients:
            per_piece_grams = float(ing['quantity_grams'])
            total_grams = per_piece_grams * total_pieces
            total_kg = total_grams / 1000
            wastage_kg = total_kg * (wastage_percent / 100)
            order_qty_kg = total_kg + wastage_kg
            cost_per_kg = float(ing['landed_cost_net_gst'] or 0)
            line_cost = order_qty_kg * cost_per_kg
            total_cost += line_cost
            
            bom_items.append({
                'ingredient_id': ing['ingredient_id'],
                'ingredient_name': ing['ingredient_name'],
                'category': ing['category_name'],
                'percentage': float(ing['percentage']),
                'per_piece_grams': round(per_piece_grams, 2),
                'total_kg': round(total_kg, 3),
                'wastage_kg': round(wastage_kg, 3),
                'order_qty_kg': round(order_qty_kg, 3),
                'cost_per_kg': round(cost_per_kg, 2),
                'line_cost': round(line_cost, 2)
            })
        
        total_weight_kg = sum(item['total_kg'] for item in bom_items)
        total_wastage_kg = sum(item['wastage_kg'] for item in bom_items)
        total_order_kg = sum(item['order_qty_kg'] for item in bom_items)
        
        cost_per_piece = total_cost / total_pieces if total_pieces > 0 else 0
        cost_per_pack = cost_per_piece * pack_count
        
        result = {
            'formulation': {
                'id': formulation['id'],
                'product_name': formulation['product_name'],
                'version': formulation['current_version'],
                'grammage': grammage
            },
            'parameters': {
                'target_quantity': target_quantity,
                'pack_count': pack_count,
                'total_pieces': total_pieces,
                'wastage_percent': wastage_percent
            },
            'items': bom_items,
            'summary': {
                'total_weight_kg': round(total_weight_kg, 3),
                'total_wastage_kg': round(total_wastage_kg, 3),
                'total_order_kg': round(total_order_kg, 3),
                'total_cost': round(total_cost, 2),
                'cost_per_piece': round(cost_per_piece, 4),
                'cost_per_pack': round(cost_per_pack, 2)
            },
            'generated_at': datetime.now().isoformat()
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500




# ============================================================================
# ESTIMATED SCORES ENDPOINT (Lather & Hardness)
# ============================================================================

@app.route('/api/formulations/<int:id>/estimated-scores', methods=['GET'])
@jwt_required()
def get_estimated_scores(id):
    """Calculate estimated lather and hardness scores from ingredient coefficients"""
    try:
        version_id = request.args.get('version_id')
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Get formulation
        cursor.execute('SELECT * FROM formulations WHERE id = ?', (id,))
        formulation = cursor.fetchone()
        
        if not formulation:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404
        
        # Get ingredients with their coefficients
        if version_id:
            cursor.execute('''
                SELECT fi.ingredient_id, fi.percentage, i.name,
                       ip.hardness_coefficient, ip.lather_coefficient,
                       ip.sap_value, ip.iodine_value, ip.ins_value
                FROM formulation_ingredients fi
                JOIN ingredients i ON fi.ingredient_id = i.id
                LEFT JOIN ingredient_properties ip ON i.id = ip.ingredient_id
                WHERE fi.formulation_id = ? AND fi.version_id = ?
            ''', (id, version_id))
        else:
            cursor.execute('''
                SELECT fi.ingredient_id, fi.percentage, i.name,
                       ip.hardness_coefficient, ip.lather_coefficient,
                       ip.sap_value, ip.iodine_value, ip.ins_value
                FROM formulation_ingredients fi
                JOIN ingredients i ON fi.ingredient_id = i.id
                LEFT JOIN ingredient_properties ip ON i.id = ip.ingredient_id
                WHERE fi.formulation_id = ?
            ''', (id,))
        
        ingredients = cursor.fetchall()
        conn.close()
        
        # Calculate weighted scores
        total_hardness = 0
        total_lather = 0
        total_percentage_with_coefficients = 0
        ingredient_contributions = []
        
        for ing in ingredients:
            percentage = float(ing['percentage']) / 100  # Convert to decimal
            hardness_coef = float(ing['hardness_coefficient'] or 0)
            lather_coef = float(ing['lather_coefficient'] or 0)
            
            if hardness_coef > 0 or lather_coef > 0:
                total_percentage_with_coefficients += float(ing['percentage'])
                total_hardness += percentage * hardness_coef
                total_lather += percentage * lather_coef
                
                ingredient_contributions.append({
                    'name': ing['name'],
                    'percentage': float(ing['percentage']),
                    'hardness_contribution': round(percentage * hardness_coef * 100, 2),
                    'lather_contribution': round(percentage * lather_coef * 100, 2),
                    'sap_value': ing['sap_value'],
                    'iodine_value': ing['iodine_value'],
                    'ins_value': ing['ins_value']
                })
        
        return jsonify({
            'formulation_id': id,
            'version_id': version_id,
            'estimated_scores': {
                'hardness': round(total_hardness * 100, 1),  # Scale to 0-100
                'lather': round(total_lather * 100, 1),      # Scale to 0-100
                'hardness_rating': 'High' if total_hardness > 0.6 else 'Medium' if total_hardness > 0.4 else 'Low',
                'lather_rating': 'High' if total_lather > 0.5 else 'Medium' if total_lather > 0.3 else 'Low'
            },
            'coverage': {
                'percentage_with_data': round(total_percentage_with_coefficients, 1),
                'note': f'{round(total_percentage_with_coefficients, 1)}% of formula has predictive data'
            },
            'ingredient_contributions': ingredient_contributions
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>/refresh-prices', methods=['POST'])
@jwt_required()
def refresh_formulation_prices(id):
    """Refresh formulation costs using current ingredient prices without changing ingredients"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Get formulation
        cursor.execute('SELECT * FROM formulations WHERE id = ?', (id,))
        formulation = cursor.fetchone()
        
        if not formulation:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404
        
        grammage = float(formulation['grammage'])
        
        # Get current ingredients
        cursor.execute('''
            SELECT fi.id, fi.ingredient_id, fi.percentage, i.landed_cost_net_gst, i.name
            FROM formulation_ingredients fi
            JOIN ingredients i ON fi.ingredient_id = i.id
            WHERE fi.formulation_id = ?
        ''', (id,))
        
        ingredients = cursor.fetchall()
        
        total_cost = 0
        updated_ingredients = []
        
        for ing in ingredients:
            percentage = float(ing['percentage'])
            cost_per_kg = float(ing['landed_cost_net_gst'] or 0)
            quantity_grams = (grammage * percentage) / 100
            cost_per_piece = (quantity_grams / 1000) * cost_per_kg
            
            # Update ingredient cost
            cursor.execute('''
                UPDATE formulation_ingredients 
                SET cost_per_piece = ?, quantity_grams = ?
                WHERE id = ?
            ''', (cost_per_piece, quantity_grams, ing['id']))
            
            total_cost += cost_per_piece
            updated_ingredients.append({
                'name': ing['name'],
                'old_cost': None,  # Could track if needed
                'new_cost': round(cost_per_piece, 4)
            })
        
        # Update formulation total cost
        cursor.execute('''
            UPDATE formulations 
            SET total_cost_per_piece = ?, updated_at = ?
            WHERE id = ?
        ''', (total_cost, datetime.now().isoformat(), id))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': 'Prices refreshed successfully',
            'new_total_cost': round(total_cost, 4),
            'updated_ingredients': updated_ingredients
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# ADVANCED SEARCH ENDPOINTS
# ============================================================================

@app.route('/api/search/ingredients', methods=['POST'])
@jwt_required()
def search_ingredients():
    """Advanced multi-criteria ingredient search"""
    try:
        data = request.get_json()
        
        conn = get_db()
        cursor = conn.cursor()
        
        query = '''
            SELECT i.*, c.name as category_name, sc.name as subcategory_name,
                   s.name as supplier_name
            FROM ingredients i
            LEFT JOIN categories c ON i.category_id = c.id
            LEFT JOIN subcategories sc ON i.subcategory_id = sc.id
            LEFT JOIN suppliers s ON i.supplier_id = s.id
            WHERE 1=1
        '''
        params = []
        
        if data.get('search'):
            query += ' AND (i.name LIKE ? OR i.inci_name LIKE ? OR i.cas_number LIKE ?)'
            search_term = f"%{data['search']}%"
            params.extend([search_term, search_term, search_term])
        
        if data.get('categories'):
            placeholders = ','.join('?' * len(data['categories']))
            query += f' AND i.category_id IN ({placeholders})'
            params.extend(data['categories'])
        
        if data.get('cost_min') is not None:
            query += ' AND i.landed_cost_net_gst >= ?'
            params.append(data['cost_min'])
        
        if data.get('cost_max') is not None:
            query += ' AND i.landed_cost_net_gst <= ?'
            params.append(data['cost_max'])
        
        if data.get('stock_status'):
            query += ' AND i.stock_status = ?'
            params.append(data['stock_status'])
        
        if data.get('supplier_id'):
            query += ' AND i.supplier_id = ?'
            params.append(data['supplier_id'])
        
        query += ' ORDER BY i.name'
        
        cursor.execute(query, params)
        results = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({
            'results': results,
            'count': len(results)
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/search/formulations', methods=['POST'])
@jwt_required()
def search_formulations():
    """Advanced multi-criteria formulation search"""
    try:
        data = request.get_json()
        
        conn = get_db()
        cursor = conn.cursor()
        
        query = '''
            SELECT DISTINCT f.*, pt.name as product_type_name,
                   COUNT(DISTINCT fi.ingredient_id) as ingredient_count
            FROM formulations f
            LEFT JOIN product_types pt ON f.product_type_id = pt.id
            LEFT JOIN formulation_ingredients fi ON f.id = fi.formulation_id
            WHERE 1=1
        '''
        params = []
        
        if data.get('search'):
            query += ' AND f.product_name LIKE ?'
            params.append(f"%{data['search']}%")
        
        if data.get('product_types'):
            placeholders = ','.join('?' * len(data['product_types']))
            query += f' AND f.product_type_id IN ({placeholders})'
            params.extend(data['product_types'])
        
        if data.get('statuses'):
            placeholders = ','.join('?' * len(data['statuses']))
            query += f' AND f.status IN ({placeholders})'
            params.extend(data['statuses'])
        
        if data.get('cost_min') is not None:
            query += ' AND f.total_cost_per_piece >= ?'
            params.append(data['cost_min'])
        
        if data.get('cost_max') is not None:
            query += ' AND f.total_cost_per_piece <= ?'
            params.append(data['cost_max'])
        
        if data.get('created_after'):
            query += ' AND f.created_at >= ?'
            params.append(data['created_after'])
        
        if data.get('created_before'):
            query += ' AND f.created_at <= ?'
            params.append(data['created_before'])
        
        if data.get('benefits'):
            placeholders = ','.join('?' * len(data['benefits']))
            query += f''' AND f.id IN (
                SELECT formulation_id FROM formulation_benefits
                WHERE benefit_id IN ({placeholders})
            )'''
            params.extend(data['benefits'])
        
        query += ' GROUP BY f.id ORDER BY f.created_at DESC'
        
        cursor.execute(query, params)
        results = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({
            'results': results,
            'count': len(results)
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# REFERENCE DATA ENDPOINTS
# ============================================================================

@app.route('/api/product-types', methods=['GET'])
@jwt_required()
def get_product_types():
    """Get all product types"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM product_types ORDER BY name')
        types = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'product_types': types}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/benefits', methods=['GET'])
@jwt_required()
def get_benefits():
    """Get all benefit categories"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM benefit_categories ORDER BY name')
        benefits = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'benefits': benefits}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/suppliers', methods=['GET'])
@jwt_required()
def get_suppliers():
    """Get all suppliers"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM suppliers ORDER BY name')
        suppliers = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'suppliers': suppliers}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tags', methods=['GET'])
@jwt_required()
def get_tags():
    """Get all tags"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM tags ORDER BY name')
        tags = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'tags': tags}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# DASHBOARD / STATISTICS ENDPOINTS
# ============================================================================

@app.route('/api/stats/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard_stats():
    """Get dashboard statistics"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) as count FROM formulations')
        total_formulations = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(*) as count FROM ingredients')
        total_ingredients = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(*) as count FROM formulations WHERE status = "active"')
        active_formulations = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(*) as count FROM formulations WHERE status = "draft"')
        draft_formulations = cursor.fetchone()['count']
        
        cursor.execute('''
            SELECT f.*, pt.name as product_type_name
            FROM formulations f
            LEFT JOIN product_types pt ON f.product_type_id = pt.id
            ORDER BY f.created_at DESC
            LIMIT 5
        ''')
        recent_formulations = [dict_from_row(row) for row in cursor.fetchall()]
        
        cursor.execute('''
            SELECT c.name, COUNT(i.id) as count
            FROM categories c
            LEFT JOIN ingredients i ON c.id = i.category_id
            GROUP BY c.id, c.name
            ORDER BY count DESC
        ''')
        category_breakdown = [dict_from_row(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({
            'totals': {
                'formulations': total_formulations,
                'ingredients': total_ingredients,
                'active_formulations': active_formulations,
                'draft_formulations': draft_formulations
            },
            'recent_formulations': recent_formulations,
            'category_breakdown': category_breakdown
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# HEALTH CHECK
# ============================================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat()
    }), 200


@app.route('/api/', methods=['GET'])
def api_root():
    """API root"""
    return jsonify({
        'message': 'Swati Soaps Formulation Management System API',
        'version': '2.1',
        'status': 'running'
    }), 200


# ============================================================================
# EXCEL FORMULATION IMPORT
# ============================================================================

@app.route('/api/formulations/import', methods=['POST'])
@jwt_required()
def import_formulations_from_excel():
    """
    Import formulations from Excel (.xlsx, .xls)

    Required columns: Ingredient Name, Percentage (must add to 100%)
    Optional columns: Supplier, HSN Code, Cost/kg
    Optional rows: Grammage, Piece per case

    When ingredient exists in DB, auto-fills: supplier, HSN, INCI, cost from DB
    When ingredient doesn't exist, creates it with provided data (or blank)
    """
    try:
        import io
        user_id = get_jwt_identity()

        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']

        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            return jsonify({'error': 'File must be Excel (.xlsx or .xls)'}), 400

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)

        conn = get_db()
        cursor = conn.cursor()

        results = {
            'sheets_processed': 0,
            'formulations_created': 0,
            'ingredients_matched': 0,
            'ingredients_created': 0,
            'errors': [],
            'enriched_ingredients': []  # Track which ingredients were auto-filled from DB
        }

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]

                formulation_data = {
                    'product_name': sheet_name,
                    'grammage': None,  # Required - must be found
                    'pack_count': 1,   # Optional, defaults to 1
                    'ingredients': []
                }

                # Find grammage (REQUIRED) and pack count (optional)
                for row in ws.iter_rows(values_only=True):
                    if row[0]:
                        cell_text = str(row[0]).lower().strip()
                        if 'grammage' in cell_text or 'gram' in cell_text or 'weight' in cell_text:
                            try:
                                formulation_data['grammage'] = float(row[1]) if row[1] else None
                            except:
                                pass
                        if 'piece' in cell_text and 'case' in cell_text:
                            try:
                                formulation_data['pack_count'] = int(row[1]) if row[1] else 1
                            except:
                                pass

                # Validate grammage is provided
                if not formulation_data['grammage']:
                    results['errors'].append(f"Sheet '{sheet_name}': Grammage is required (add row with 'Grammage' label and value)")
                    continue

                # Find ingredients section - look for header row
                ingredient_start_row = None
                for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row[0]:
                        cell_text = str(row[0]).lower().strip()
                        # Accept various header names
                        if cell_text in ['particulars', 'ingredient', 'ingredients', 'name', 'item']:
                            ingredient_start_row = idx + 1
                            break

                if not ingredient_start_row:
                    results['errors'].append(f"Sheet '{sheet_name}': No ingredient header found (use 'Particulars' or 'Ingredient')")
                    continue

                # Extract ingredients - only name and percentage are required
                total_percentage = 0
                for row in ws.iter_rows(min_row=ingredient_start_row, values_only=True):
                    # Get ingredient name (Column A) - REQUIRED
                    ingredient_name = row[0] if len(row) > 0 else None
                    if not ingredient_name:
                        break

                    ingredient_name = str(ingredient_name).strip()

                    # Skip total row
                    if 'total' in ingredient_name.lower():
                        break

                    # Get percentage (Column C or B) - REQUIRED
                    percentage = None
                    if len(row) > 2 and row[2]:  # Column C
                        percentage = row[2]
                    elif len(row) > 1 and row[1]:  # Column B as fallback
                        try:
                            percentage = float(row[1])
                        except:
                            pass

                    if percentage is None:
                        results['errors'].append(f"Sheet '{sheet_name}': Missing percentage for '{ingredient_name}'")
                        continue

                    # Convert percentage - handle both 94 and 0.94 formats
                    try:
                        pct_value = float(percentage)
                        # If value is <= 1 and not 0 or 1 exactly, assume it's decimal format
                        if pct_value > 0 and pct_value < 1:
                            pct_value = pct_value * 100
                        # Now pct_value should be in 0-100 range
                    except:
                        results['errors'].append(f"Sheet '{sheet_name}': Invalid percentage '{percentage}' for '{ingredient_name}'")
                        continue

                    total_percentage += pct_value

                    # Optional fields from Excel
                    excel_supplier = str(row[1]).strip() if len(row) > 1 and row[1] else None
                    excel_hsn = str(row[3]).strip() if len(row) > 3 and row[3] else None
                    excel_cost = float(row[5]) if len(row) > 5 and row[5] else 0

                    # Search for ingredient in database
                    cursor.execute('''
                        SELECT i.id, i.name, i.landed_cost_net_gst, i.hsn_code, i.inci_name,
                               s.name as supplier_name, c.name as category_name
                        FROM ingredients i
                        LEFT JOIN suppliers s ON i.supplier_id = s.id
                        LEFT JOIN categories c ON i.category_id = c.id
                        WHERE LOWER(i.name) = LOWER(?)
                    ''', (ingredient_name,))

                    ingredient_match = cursor.fetchone()

                    if ingredient_match:
                        # Ingredient found - use DB values (auto-fill)
                        ingredient_id = ingredient_match[0]
                        db_name = ingredient_match[1]
                        unit_cost = ingredient_match[2] or 0
                        db_hsn = ingredient_match[3]
                        db_inci = ingredient_match[4]
                        db_supplier = ingredient_match[5]
                        db_category = ingredient_match[6]

                        results['ingredients_matched'] += 1
                        results['enriched_ingredients'].append({
                            'name': db_name,
                            'from_db': True,
                            'supplier': db_supplier,
                            'hsn': db_hsn,
                            'inci': db_inci,
                            'cost': unit_cost
                        })
                    else:
                        # Ingredient not found - create new with Excel data (or blank)
                        try:
                            supplier_id = None
                            if excel_supplier:
                                cursor.execute('SELECT id FROM suppliers WHERE LOWER(name) = LOWER(?)', (excel_supplier,))
                                supplier_result = cursor.fetchone()
                                if supplier_result:
                                    supplier_id = supplier_result[0]
                                else:
                                    cursor.execute('INSERT INTO suppliers (name, created_at) VALUES (?, ?)',
                                                 (excel_supplier, datetime.now().isoformat()))
                                    supplier_id = cursor.lastrowid

                            # Get or create Uncategorized category
                            cursor.execute('SELECT id FROM categories WHERE name = ? LIMIT 1', ('Uncategorized',))
                            category_result = cursor.fetchone()
                            if not category_result:
                                cursor.execute('INSERT INTO categories (name, created_at) VALUES (?, ?)',
                                             ('Uncategorized', datetime.now().isoformat()))
                                category_id = cursor.lastrowid
                            else:
                                category_id = category_result[0]

                            cursor.execute('''
                                INSERT INTO ingredients (
                                    name, category_id, supplier_id, landed_cost_net_gst,
                                    hsn_code, stock_status, unit_of_measure, created_at, updated_at
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                ingredient_name,
                                category_id,
                                supplier_id,
                                excel_cost,
                                excel_hsn,
                                'in_stock',
                                'kg',
                                datetime.now().isoformat(),
                                datetime.now().isoformat()
                            ))

                            ingredient_id = cursor.lastrowid
                            unit_cost = excel_cost
                            results['ingredients_created'] += 1

                        except Exception as e:
                            results['errors'].append(f"Failed to create '{ingredient_name}': {str(e)}")
                            continue

                    formulation_data['ingredients'].append({
                        'ingredient_id': ingredient_id,
                        'ingredient_name': ingredient_name,
                        'percentage': pct_value,
                        'unit_cost': unit_cost
                    })

                # Validate total percentage (must be ~100%)
                if abs(total_percentage - 100) > 0.5:
                    results['errors'].append(f"Sheet '{sheet_name}': Percentages sum to {total_percentage:.2f}% (must be 100%)")
                    continue

                # Create formulation if we have ingredients
                if len(formulation_data['ingredients']) > 0:
                    cursor.execute('SELECT id FROM product_types WHERE name LIKE ? LIMIT 1', ('%Soap%',))
                    product_type_result = cursor.fetchone()
                    product_type_id = product_type_result[0] if product_type_result else 1

                    cursor.execute('''
                        INSERT INTO formulations (
                            product_name, product_type_id, grammage, pack_count,
                            status, current_version, created_by, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        formulation_data['product_name'],
                        product_type_id,
                        formulation_data['grammage'],
                        formulation_data['pack_count'],
                        'draft',
                        'v1.0',
                        user_id,
                        datetime.now().isoformat(),
                        datetime.now().isoformat()
                    ))

                    formulation_id = cursor.lastrowid

                    # Create version snapshot
                    snapshot = {
                        'grammage': formulation_data['grammage'],
                        'pack_count': formulation_data['pack_count'],
                        'ingredients': formulation_data['ingredients']
                    }

                    cursor.execute('''
                        INSERT INTO formulation_versions (
                            formulation_id, version_number, created_at, created_by,
                            change_notes, ingredients_snapshot, cost_snapshot
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        formulation_id,
                        'v1.0',
                        datetime.now().isoformat(),
                        user_id,
                        'Imported from Excel',
                        json.dumps(snapshot),
                        0
                    ))

                    version_id = cursor.lastrowid

                    # Insert formulation ingredients and calculate cost
                    total_cost = 0
                    for ing in formulation_data['ingredients']:
                        quantity_grams = (ing['percentage'] / 100) * formulation_data['grammage']
                        cost_per_piece = (quantity_grams / 1000) * ing['unit_cost']
                        total_cost += cost_per_piece

                        cursor.execute('''
                            INSERT INTO formulation_ingredients (
                                formulation_id, version_id, ingredient_id, percentage,
                                quantity_grams, cost_per_piece
                            ) VALUES (?, ?, ?, ?, ?, ?)
                        ''', (
                            formulation_id,
                            version_id,
                            ing['ingredient_id'],
                            ing['percentage'],
                            quantity_grams,
                            cost_per_piece
                        ))

                    # Update formulation total cost
                    cursor.execute('UPDATE formulations SET total_cost_per_piece = ? WHERE id = ?',
                                 (round(total_cost, 4), formulation_id))

                    results['formulations_created'] += 1

                results['sheets_processed'] += 1

            except Exception as e:
                results['errors'].append(f"Sheet '{sheet_name}': {str(e)}")

        conn.commit()
        conn.close()

        # Calculate skipped
        skipped = results['sheets_processed'] - results['formulations_created'] + len(results['errors'])

        return jsonify({
            'message': 'Import completed',
            'imported': results['formulations_created'],
            'skipped': skipped,
            'sheets_processed': results['sheets_processed'],
            'formulations_created': results['formulations_created'],
            'ingredients_matched': results['ingredients_matched'],
            'ingredients_created': results['ingredients_created'],
            'errors': results['errors']
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# ============================================================================
# ROLE PERMISSION DECORATOR (Add after imports)
# ============================================================================

def role_required(*allowed_roles):
    """Decorator to check user role permissions"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user_id = get_jwt_identity()
            
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT role FROM users WHERE id = ?', (user_id,))
            user = cursor.fetchone()
            conn.close()
            
            if not user or user['role'] not in allowed_roles:
                return jsonify({'error': 'Permission denied. Required roles: ' + ', '.join(allowed_roles)}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ============================================================================
# ADMIN USER MANAGEMENT ENDPOINTS
# ============================================================================

@app.route('/api/admin/users', methods=['GET'])
@jwt_required()
@role_required('owner', 'admin')
def get_all_users():
    """Get all users (admin only)"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, email, full_name, role, is_active, last_login, created_at
            FROM users
            ORDER BY role, full_name
        ''')
        
        users = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'users': users}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users', methods=['POST'])
@jwt_required()
@role_required('owner', 'admin')
def create_user():
    """Create new user (admin only)"""
    try:
        data = request.get_json()
        
        email = data.get('email')
        password = data.get('password')
        full_name = data.get('full_name', '')
        role = data.get('role', 'viewer')
        is_active = data.get('is_active', True)
        
        if not email or not password:
            return jsonify({'error': 'Email and password are required'}), 400
        
        # Validate role
        valid_roles = ['viewer', 'qc', 'accountant', 'owner', 'admin']
        if role not in valid_roles:
            return jsonify({'error': f'Invalid role. Must be one of: {", ".join(valid_roles)}'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if email already exists
        cursor.execute('SELECT id FROM users WHERE email = ?', (email,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Email already exists'}), 400
        
        # Insert new user (note: in production, use proper password hashing!)
        cursor.execute('''
            INSERT INTO users (email, password_hash, full_name, role, is_active, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (email, password, full_name, role, is_active, datetime.now().isoformat()))
        
        user_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': 'User created successfully',
            'user_id': user_id
        }), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@jwt_required()
@role_required('owner', 'admin')
def update_user(user_id):
    """Update user (admin only)"""
    try:
        data = request.get_json()
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check user exists
        cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        
        if not user:
            conn.close()
            return jsonify({'error': 'User not found'}), 404
        
        # Build update query dynamically
        updates = []
        params = []
        
        if 'email' in data:
            updates.append('email = ?')
            params.append(data['email'])
        
        if 'full_name' in data:
            updates.append('full_name = ?')
            params.append(data['full_name'])
        
        if 'password' in data and data['password']:
            updates.append('password_hash = ?')
            params.append(data['password'])  # In production, hash this!
        
        if 'role' in data:
            valid_roles = ['viewer', 'qc', 'accountant', 'owner', 'admin']
            if data['role'] not in valid_roles:
                conn.close()
                return jsonify({'error': f'Invalid role'}), 400
            updates.append('role = ?')
            params.append(data['role'])
        
        if 'is_active' in data:
            updates.append('is_active = ?')
            params.append(1 if data['is_active'] else 0)
        
        if updates:
            params.append(user_id)
            cursor.execute(f'''
                UPDATE users SET {', '.join(updates)}
                WHERE id = ?
            ''', params)
            conn.commit()
        
        conn.close()
        
        return jsonify({'message': 'User updated successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
@role_required('admin')
def delete_user(user_id):
    """Delete user (admin only)"""
    try:
        current_user_id = get_jwt_identity()
        
        if int(current_user_id) == user_id:
            return jsonify({'error': 'Cannot delete your own account'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'error': 'User not found'}), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'User deleted successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# APPROVAL WORKFLOW ENDPOINTS
# ============================================================================

@app.route('/api/admin/pending-approvals', methods=['GET'])
@jwt_required()
@role_required('owner', 'admin')
def get_pending_approvals():
    """Get all pending approval requests"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                fv.id as version_id,
                fv.formulation_id,
                fv.version_number,
                fv.change_notes,
                fv.approval_status,
                fv.submitted_at,
                fv.submitted_by,
                f.product_name,
                u.full_name as submitted_by_name
            FROM formulation_versions fv
            JOIN formulations f ON fv.formulation_id = f.id
            LEFT JOIN users u ON fv.submitted_by = u.id
            WHERE fv.approval_status = 'pending'
            ORDER BY fv.submitted_at ASC
        ''')
        
        approvals = [dict_from_row(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({'pending_approvals': approvals}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/approvals/<int:version_id>/approve', methods=['POST'])
@jwt_required()
@role_required('owner', 'admin')
def approve_version(version_id):
    """Approve a formulation version"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json() or {}
        notes = data.get('notes', '')
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check version exists and is pending
        cursor.execute('''
            SELECT * FROM formulation_versions WHERE id = ? AND approval_status = 'pending'
        ''', (version_id,))
        version = cursor.fetchone()
        
        if not version:
            conn.close()
            return jsonify({'error': 'Version not found or not pending approval'}), 404
        
        # Approve the version
        cursor.execute('''
            UPDATE formulation_versions
            SET approval_status = 'approved',
                approved_by = ?,
                approved_at = ?,
                approval_notes = ?
            WHERE id = ?
        ''', (user_id, datetime.now().isoformat(), notes, version_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Version approved successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/approvals/<int:version_id>/reject', methods=['POST'])
@jwt_required()
@role_required('owner', 'admin')
def reject_version(version_id):
    """Reject a formulation version"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json() or {}
        notes = data.get('notes', '')
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check version exists and is pending
        cursor.execute('''
            SELECT * FROM formulation_versions WHERE id = ? AND approval_status = 'pending'
        ''', (version_id,))
        version = cursor.fetchone()
        
        if not version:
            conn.close()
            return jsonify({'error': 'Version not found or not pending approval'}), 404
        
        # Reject the version
        cursor.execute('''
            UPDATE formulation_versions
            SET approval_status = 'rejected',
                approved_by = ?,
                approved_at = ?,
                approval_notes = ?
            WHERE id = ?
        ''', (user_id, datetime.now().isoformat(), notes, version_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Version rejected'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/formulations/<int:id>/submit-for-approval', methods=['POST'])
@jwt_required()
def submit_for_approval(id):
    """Submit a formulation version for owner approval"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json() or {}
        version_id = data.get('version_id')
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Get the version to submit (latest if not specified)
        if version_id:
            cursor.execute('''
                SELECT * FROM formulation_versions
                WHERE id = ? AND formulation_id = ?
            ''', (version_id, id))
        else:
            cursor.execute('''
                SELECT * FROM formulation_versions
                WHERE formulation_id = ?
                ORDER BY created_at DESC LIMIT 1
            ''', (id,))
        
        version = cursor.fetchone()
        
        if not version:
            conn.close()
            return jsonify({'error': 'Version not found'}), 404
        
        # Update to pending status
        cursor.execute('''
            UPDATE formulation_versions
            SET approval_status = 'pending',
                submitted_by = ?,
                submitted_at = ?
            WHERE id = ?
        ''', (user_id, datetime.now().isoformat(), version['id']))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Submitted for approval'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# MAKE FORMULATION ACTIVE (Owner only)
# ============================================================================

@app.route('/api/formulations/<int:id>/make-active', methods=['POST'])
@jwt_required()
@role_required('owner', 'admin')
def make_formulation_active(id):
    """Make a formulation active (owner only)"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Check formulation exists
        cursor.execute('SELECT * FROM formulations WHERE id = ?', (id,))
        formulation = cursor.fetchone()
        
        if not formulation:
            conn.close()
            return jsonify({'error': 'Formulation not found'}), 404
        
        # Check if latest version is rejected
        cursor.execute('''
            SELECT approval_status FROM formulation_versions
            WHERE formulation_id = ?
            ORDER BY created_at DESC LIMIT 1
        ''', (id,))
        latest_version = cursor.fetchone()
        
        if latest_version and latest_version['approval_status'] == 'rejected':
            conn.close()
            return jsonify({
                'error': 'Cannot make active: latest version is rejected. It must be approved first.'
            }), 400
        
        # Update status to active
        cursor.execute('''
            UPDATE formulations
            SET status = 'active', updated_at = ?
            WHERE id = ?
        ''', (datetime.now().isoformat(), id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Formulation is now active'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# DATABASE MIGRATION FOR APPROVAL FIELDS
# Run this SQL on your database to add the approval workflow fields:
# ============================================================================
# RUN APPLICATION
# ============================================================================

if __name__ == '__main__':
    if not os.path.exists(app.config['DATABASE']):
        print("⚠️  WARNING: Database file not found!")
        print(f"   Expected location: {app.config['DATABASE']}")
    else:
        print(f"✅ Database found: {app.config['DATABASE']}")
    
    print("\n" + "="*70)
    print("  SWATI SOAPS FORMULATION MANAGEMENT SYSTEM")
    print("  Backend API Server v2.1 (Production-Ready)")
    print("="*70)
    print(f"\n🚀 Starting server on http://localhost:5000")
    print(f"📊 API endpoints available at http://localhost:5000/api/")
    print(f"🔒 CORS enabled for: http://localhost:3000")
    print(f"\n💡 Press Ctrl+C to stop the server\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
