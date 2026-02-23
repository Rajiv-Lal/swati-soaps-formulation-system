#!/usr/bin/env python3
"""
Test script for formulation Excel import
"""

import os
import sys
import io
import json
import sqlite3
from datetime import datetime

# Add backend to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl import Workbook

def create_test_excel():
    """Create a test Excel file with sample formulation"""
    wb = Workbook()

    # Sheet 1: Valid formulation
    ws1 = wb.active
    ws1.title = "Test Moisturizing Soap"

    # Grammage (required)
    ws1['A1'] = 'Grammage'
    ws1['B1'] = 75

    # Piece per case (optional)
    ws1['A2'] = 'Piece per case'
    ws1['B2'] = 3

    # Empty row
    ws1['A3'] = ''

    # Ingredients header
    ws1['A4'] = 'Ingredient'
    ws1['B4'] = 'Supplier'
    ws1['C4'] = '%'
    ws1['D4'] = 'HSN'
    ws1['E4'] = ''
    ws1['F4'] = 'Cost/kg'

    # Ingredients data (only name and % filled, rest optional)
    ingredients = [
        ('Soap Noodles', '', 92, '', '', ''),
        ('Niacinamide', '', 3, '', '', ''),
        ('Glycerin', '', 2, '', '', ''),
        ('Fragrance Oil', '', 2, '', '', ''),
        ('Titanium Dioxide', '', 1, '', '', ''),
    ]

    for i, (name, supplier, pct, hsn, _, cost) in enumerate(ingredients, start=5):
        ws1[f'A{i}'] = name
        ws1[f'B{i}'] = supplier
        ws1[f'C{i}'] = pct
        ws1[f'D{i}'] = hsn
        ws1[f'F{i}'] = cost

    # Total row
    ws1['A10'] = 'Total'
    ws1['C10'] = 100

    # Sheet 2: Another valid formulation with decimal percentages
    ws2 = wb.create_sheet("Test Brightening Soap")
    ws2['A1'] = 'Grammage'
    ws2['B1'] = 100

    ws2['A3'] = 'Particulars'  # Alternative header name
    ws2['C3'] = '%'

    # Using decimal format (0.95 = 95%)
    ws2['A4'] = 'Soap Noodles'
    ws2['C4'] = 0.95
    ws2['A5'] = 'Alpha Arbutin'
    ws2['C5'] = 0.03
    ws2['A6'] = 'Vitamin E'
    ws2['C6'] = 0.02
    ws2['A7'] = 'Total'
    ws2['C7'] = 1.0

    # Sheet 3: Invalid - missing grammage
    ws3 = wb.create_sheet("Missing Grammage Test")
    ws3['A1'] = 'Ingredient'
    ws3['C1'] = '%'
    ws3['A2'] = 'Soap Noodles'
    ws3['C2'] = 100

    # Sheet 4: Invalid - percentages don't add to 100
    ws4 = wb.create_sheet("Wrong Percentage Test")
    ws4['A1'] = 'Grammage'
    ws4['B1'] = 75
    ws4['A3'] = 'Ingredient'
    ws4['C3'] = '%'
    ws4['A4'] = 'Soap Noodles'
    ws4['C4'] = 50  # Only 50%, should fail

    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Also save to file for inspection
    test_file_path = os.path.join(os.path.dirname(__file__), 'test_formulation_import.xlsx')
    wb.save(test_file_path)
    print(f"✅ Test Excel file created: {test_file_path}")

    return excel_buffer, test_file_path


def test_import_directly():
    """Test the import logic directly without HTTP"""
    print("\n" + "="*60)
    print("FORMULATION IMPORT TEST")
    print("="*60)

    # Create test Excel
    excel_buffer, test_file_path = create_test_excel()

    # Load database
    db_path = os.path.join(os.path.dirname(__file__), 'swati_soaps.db')
    if not os.path.exists(db_path):
        print(f"❌ Database not found: {db_path}")
        return False

    print(f"✅ Database found: {db_path}")

    # Check existing ingredients
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # Check for some test ingredients
    test_ingredients = ['Soap Noodles', 'Niacinamide', 'Glycerin', 'Alpha Arbutin', 'Vitamin E']
    print("\n📦 Checking existing ingredients:")

    for ing_name in test_ingredients:
        cursor.execute('''
            SELECT i.id, i.name, i.landed_cost_net_gst, i.hsn_code,
                   s.name as supplier_name
            FROM ingredients i
            LEFT JOIN suppliers s ON i.supplier_id = s.id
            WHERE LOWER(i.name) LIKE LOWER(?)
        ''', (f'%{ing_name}%',))
        result = cursor.fetchone()
        if result:
            print(f"  ✅ {ing_name}: Found (ID={result['id']}, Cost={result['landed_cost_net_gst']}, Supplier={result['supplier_name']})")
        else:
            print(f"  ❌ {ing_name}: NOT FOUND (will be created)")

    conn.close()

    # Now test the actual import
    print("\n🔄 Testing import logic...")

    import openpyxl as xl
    wb = xl.load_workbook(excel_buffer, data_only=True)

    print(f"\n📑 Sheets found: {wb.sheetnames}")

    results = {
        'sheets_processed': 0,
        'formulations_valid': 0,
        'formulations_invalid': 0,
        'errors': []
    }

    for sheet_name in wb.sheetnames:
        print(f"\n--- Processing: {sheet_name} ---")
        ws = wb[sheet_name]

        # Find grammage
        grammage = None
        pack_count = 1

        for row in ws.iter_rows(values_only=True):
            if row[0]:
                cell_text = str(row[0]).lower().strip()
                if 'grammage' in cell_text or 'gram' in cell_text:
                    try:
                        grammage = float(row[1]) if row[1] else None
                        print(f"  Grammage: {grammage}")
                    except:
                        pass
                if 'piece' in cell_text and 'case' in cell_text:
                    try:
                        pack_count = int(row[1]) if row[1] else 1
                        print(f"  Pack count: {pack_count}")
                    except:
                        pass

        if not grammage:
            print(f"  ❌ ERROR: Grammage is required")
            results['errors'].append(f"{sheet_name}: Missing grammage")
            results['formulations_invalid'] += 1
            continue

        # Find ingredients header
        ingredient_start_row = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row[0]:
                cell_text = str(row[0]).lower().strip()
                if cell_text in ['particulars', 'ingredient', 'ingredients', 'name', 'item']:
                    ingredient_start_row = idx + 1
                    print(f"  Ingredients start at row: {ingredient_start_row}")
                    break

        if not ingredient_start_row:
            print(f"  ❌ ERROR: No ingredient header found")
            results['errors'].append(f"{sheet_name}: No ingredient header")
            results['formulations_invalid'] += 1
            continue

        # Extract ingredients
        total_pct = 0
        ingredients = []

        for row in ws.iter_rows(min_row=ingredient_start_row, values_only=True):
            name = row[0] if len(row) > 0 else None
            if not name:
                break
            name = str(name).strip()
            if 'total' in name.lower():
                break

            # Get percentage
            pct = None
            if len(row) > 2 and row[2]:
                pct = row[2]
            elif len(row) > 1 and row[1]:
                try:
                    pct = float(row[1])
                except:
                    pass

            if pct is None:
                print(f"  ⚠️ Missing % for: {name}")
                continue

            pct_value = float(pct)
            # Handle decimal format
            if pct_value > 0 and pct_value < 1:
                pct_value = pct_value * 100

            total_pct += pct_value
            ingredients.append({'name': name, 'percentage': pct_value})
            print(f"  📦 {name}: {pct_value}%")

        print(f"  Total: {total_pct}%")

        # Validate 100%
        if abs(total_pct - 100) > 0.5:
            print(f"  ❌ ERROR: Percentages sum to {total_pct}% (must be 100%)")
            results['errors'].append(f"{sheet_name}: Sum is {total_pct}%")
            results['formulations_invalid'] += 1
            continue

        print(f"  ✅ VALID: {len(ingredients)} ingredients, {grammage}g")
        results['formulations_valid'] += 1
        results['sheets_processed'] += 1

    # Summary
    print("\n" + "="*60)
    print("TEST RESULTS")
    print("="*60)
    print(f"Sheets processed: {results['sheets_processed']}")
    print(f"Valid formulations: {results['formulations_valid']}")
    print(f"Invalid formulations: {results['formulations_invalid']}")
    if results['errors']:
        print(f"Errors:")
        for err in results['errors']:
            print(f"  - {err}")

    expected_valid = 2  # Sheet 1 and 2
    expected_invalid = 2  # Sheet 3 (no grammage) and 4 (wrong %)

    if results['formulations_valid'] == expected_valid and results['formulations_invalid'] == expected_invalid:
        print(f"\n✅ ALL TESTS PASSED!")
        return True
    else:
        print(f"\n❌ TEST FAILED: Expected {expected_valid} valid, {expected_invalid} invalid")
        return False


if __name__ == '__main__':
    success = test_import_directly()
    sys.exit(0 if success else 1)
