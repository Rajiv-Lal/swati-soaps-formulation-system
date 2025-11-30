#!/usr/bin/env python3
"""
UNIFIED INGREDIENT IMPORT
1. Deduplicate DB + Excel into single master list
2. Insert ALL into database (with whatever data available)
3. Generate enrichment template for human editing
"""

import sqlite3
import json
import openpyxl
from datetime import datetime

DB_PATH = 'swati_soaps.db'
EXCEL_PATH = '/mnt/user-data/uploads/Globalbees_75_gram_soap_costing_June_2025.xlsx'

def load_existing_ingredients():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM ingredients ORDER BY id')
    existing = [dict(row) for row in cursor.fetchall()]
    conn.close()
    existing_lookup = {ing['name'].lower().strip(): ing for ing in existing}
    print(f"✅ Loaded {len(existing)} existing ingredients from DB")
    return existing_lookup

def extract_from_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    excel_ingredients = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ingredient_start = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row[0] and 'particulars' in str(row[0]).lower():
                ingredient_start = idx + 1
                break
        if ingredient_start:
            for row in ws.iter_rows(min_row=ingredient_start, values_only=True):
                name = row[0]
                supplier = row[1]
                hsn_code = row[3]
                cost = row[5]
                if not name or 'total' in str(name).lower():
                    break
                key = str(name).strip().lower()
                if key not in excel_ingredients:
                    excel_ingredients[key] = {
                        'name': str(name).strip(),
                        'supplier_name': str(supplier).strip() if supplier else None,
                        'hsn_code': str(hsn_code).strip() if hsn_code else None,
                        'cost': float(cost) if cost else None,
                        'usage_count': 0,
                        'sheets': []
                    }
                excel_ingredients[key]['usage_count'] += 1
                excel_ingredients[key]['sheets'].append(sheet_name)
    print(f"✅ Extracted {len(excel_ingredients)} unique ingredients from Excel")
    return excel_ingredients

def create_unified_list(existing_lookup, excel_ingredients):
    unified = []
    all_keys = set(existing_lookup.keys()) | set(excel_ingredients.keys())
    for key in all_keys:
        in_db = key in existing_lookup
        in_excel = key in excel_ingredients
        entry = {'name': '', 'source': '', 'db_id': None, 'db_data': {}, 'excel_data': {}, 'needs_enrichment': False, 'action': ''}
        if in_db and in_excel:
            entry['name'] = existing_lookup[key]['name']
            entry['source'] = 'BOTH'
            entry['db_id'] = existing_lookup[key]['id']
            entry['db_data'] = existing_lookup[key]
            entry['excel_data'] = excel_ingredients[key]
            entry['needs_enrichment'] = not (existing_lookup[key].get('inci_name') and existing_lookup[key].get('cas_number'))
            entry['action'] = 'UPDATE_FROM_EXCEL'
        elif in_db:
            entry['name'] = existing_lookup[key]['name']
            entry['source'] = 'DB_ONLY'
            entry['db_id'] = existing_lookup[key]['id']
            entry['db_data'] = existing_lookup[key]
            entry['needs_enrichment'] = not (existing_lookup[key].get('inci_name') and existing_lookup[key].get('cas_number'))
            entry['action'] = 'KEEP'
        else:
            entry['name'] = excel_ingredients[key]['name']
            entry['source'] = 'EXCEL_ONLY'
            entry['excel_data'] = excel_ingredients[key]
            entry['needs_enrichment'] = True
            entry['action'] = 'INSERT_NEW'
        unified.append(entry)
    unified.sort(key=lambda x: x['excel_data'].get('usage_count', 0), reverse=True)
    print(f"\n{'='*70}\nUNIFIED MASTER LIST\n{'='*70}")
    print(f"Total unique ingredients: {len(unified)}")
    print(f"  - DB only:              {sum(1 for x in unified if x['source'] == 'DB_ONLY')}")
    print(f"  - Excel only (NEW):     {sum(1 for x in unified if x['source'] == 'EXCEL_ONLY')}")
    print(f"  - Both (existing):      {sum(1 for x in unified if x['source'] == 'BOTH')}")
    print(f"  - Need enrichment:      {sum(1 for x in unified if x['needs_enrichment'])}")
    print(f"{'='*70}\n")
    return unified

def bulk_insert_update(unified):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM categories WHERE name = 'Uncategorized' LIMIT 1")
    result = cursor.fetchone()
    if result:
        uncategorized_id = result[0]
    else:
        cursor.execute("INSERT INTO categories (name, created_at) VALUES ('Uncategorized', ?)", (datetime.now().isoformat(),))
        uncategorized_id = cursor.lastrowid
    inserted = 0
    updated = 0
    for entry in unified:
        if entry['action'] == 'INSERT_NEW':
            supplier_id = None
            if entry['excel_data'].get('supplier_name'):
                cursor.execute("SELECT id FROM suppliers WHERE LOWER(name) = LOWER(?)", (entry['excel_data']['supplier_name'],))
                result = cursor.fetchone()
                if result:
                    supplier_id = result[0]
                else:
                    cursor.execute("INSERT INTO suppliers (name, created_at) VALUES (?, ?)", (entry['excel_data']['supplier_name'], datetime.now().isoformat()))
                    supplier_id = cursor.lastrowid
            cursor.execute('''
                INSERT INTO ingredients (
                    name, category_id, supplier_id, landed_cost_net_gst,
                    hsn_code, stock_status, unit_of_measure,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (entry['name'], uncategorized_id, supplier_id, entry['excel_data'].get('cost', 0), entry['excel_data'].get('hsn_code'), 'in_stock', 'kg', datetime.now().isoformat(), datetime.now().isoformat()))
            entry['db_id'] = cursor.lastrowid
            inserted += 1
        elif entry['action'] == 'UPDATE_FROM_EXCEL':
            supplier_id = entry['db_data'].get('supplier_id')
            if entry['excel_data'].get('supplier_name') and not supplier_id:
                cursor.execute("SELECT id FROM suppliers WHERE LOWER(name) = LOWER(?)", (entry['excel_data']['supplier_name'],))
                result = cursor.fetchone()
                if result:
                    supplier_id = result[0]
                else:
                    cursor.execute("INSERT INTO suppliers (name, created_at) VALUES (?, ?)", (entry['excel_data']['supplier_name'], datetime.now().isoformat()))
                    supplier_id = cursor.lastrowid
            cursor.execute('''
                UPDATE ingredients SET
                    landed_cost_net_gst = ?,
                    supplier_id = ?,
                    hsn_code = COALESCE(hsn_code, ?),
                    updated_at = ?
                WHERE id = ?
            ''', (entry['excel_data'].get('cost', entry['db_data'].get('landed_cost_net_gst')), supplier_id, entry['excel_data'].get('hsn_code'), datetime.now().isoformat(), entry['db_id']))
            updated += 1
    conn.commit()
    conn.close()
    print(f"✅ Database updated:\n   - Inserted: {inserted} new ingredients\n   - Updated:  {updated} existing ingredients")
    return unified

def generate_enrichment_template(unified):
    need_enrichment = [x for x in unified if x['needs_enrichment']]
    need_enrichment.sort(key=lambda x: x['excel_data'].get('usage_count', 0), reverse=True)
    template = []
    for entry in need_enrichment:
        template.append({
            'db_id': entry['db_id'],
            'name': entry['name'],
            'current_data': {
                'supplier': entry['excel_data'].get('supplier_name') or entry['db_data'].get('supplier_id'),
                'cost': entry['excel_data'].get('cost') or entry['db_data'].get('landed_cost_net_gst'),
                'hsn_code': entry['excel_data'].get('hsn_code') or entry['db_data'].get('hsn_code'),
                'inci_name': entry['db_data'].get('inci_name'),
                'cas_number': entry['db_data'].get('cas_number'),
                'category_id': entry['db_data'].get('category_id')
            },
            'usage_count': entry['excel_data'].get('usage_count', 0),
            'priority': 'HIGH' if entry['excel_data'].get('usage_count', 0) >= 10 else 'MEDIUM' if entry['excel_data'].get('usage_count', 0) >= 5 else 'LOW',
            'to_enrich': {'inci_name': '', 'cas_number': '', 'category': '', 'description': '', 'storage_conditions': '', 'shelf_life_months': None, 'usage_rate_min': None, 'usage_rate_max': None, 'tags': [], 'source_url': ''}
        })
    with open('/tmp/enrichment_template.json', 'w') as f:
        json.dump(template, f, indent=2)
    print(f"\n✅ Enrichment template created:\n   📁 /tmp/enrichment_template.json\n   📋 {len(template)} ingredients need enrichment")
    with open('/tmp/unified_ingredients.json', 'w') as f:
        json.dump(unified, f, indent=2, default=str)
    print(f"   📁 /tmp/unified_ingredients.json (complete master list)")

if __name__ == "__main__":
    print("\n" + "="*70 + "\nUNIFIED INGREDIENT IMPORT\n" + "="*70 + "\n")
    existing = load_existing_ingredients()
    excel = extract_from_excel()
    unified = create_unified_list(existing, excel)
    unified = bulk_insert_update(unified)
    generate_enrichment_template(unified)
    print("\n" + "="*70 + "\n✅ ALL INGREDIENTS NOW IN DATABASE\n" + "="*70 + "\n\nNEXT STEPS:\n1. Edit /tmp/enrichment_template.json with public data\n2. Run enrichment update script\n3. Import formulations\n" + "="*70 + "\n")
